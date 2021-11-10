from pathlib import Path
import pandas as pd
import openpyxl

class Output:

    def clean_data(self, scrape_data): 
        print('Cleaning data..')
        scrape_data.drop(['author_pub_id', 'source', 'url_related_articles', 'citedby_url', 'cites_id', 'cites_per_year', 'container_type', 'num_citations', 'filled'], axis=1, inplace=True)
        cleaned_data = pd.DataFrame()
        for bib in scrape_data['bib']:
            cleaned_data = cleaned_data.append(bib, ignore_index=True)
        cleaned_data['pub_url'] = scrape_data['pub_url']
        return cleaned_data

    def xlsoutput(self, data, path):
        print('Writing data into xlsx..')
        pub_xlsx = Path(path)
        wb_obj = openpyxl.load_workbook(pub_xlsx)
        sheet = wb_obj.active
        w_row = sheet.max_row + 1

        # Loop through publications of authors
        for index, row in data.iterrows():
            sheet.cell(row=w_row, column=1).value = row['title']
            sheet.cell(row=w_row, column=2).value = row['author']
            sheet.cell(row=w_row, column=3).value = row['pub_url']
            sheet.cell(row=w_row, column=4).value = row['pub_year']
            w_row+=1
        wb_obj.save(filename=path)
