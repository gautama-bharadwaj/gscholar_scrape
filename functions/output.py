from pathlib import Path
import pandas as pd
import math
import openpyxl
from functions.email_notif import Email
from openpyxl.styles import Alignment


class Output:

    def clean_data(self, scrape_data):
        print('Cleaning data..')
        scrape_data.drop(['author_pub_id', 'source', 'url_related_articles', 'citedby_url', 'cites_id',
                          'cites_per_year', 'container_type', 'num_citations', 'filled'], axis=1, inplace=True)
        cleaned_data = pd.DataFrame()
        for bib in scrape_data['bib']:
            cleaned_data = cleaned_data.append(bib, ignore_index=True)
        cleaned_data['pub_url'] = scrape_data['pub_url']
        return cleaned_data

    def xlsoutput(self, data, path):
        print('Writing data into xlsx..')
        pub_xlsx = Path(path)
        wb_obj = openpyxl.load_workbook(pub_xlsx)
        sheet = wb_obj.active
        i = 1
        existing_titles = []
        while True:
            if (sheet.cell(row=i, column=3).value is None and sheet.cell(row=i, column=4).value is None and sheet.cell(row=i, column=6).value is None and sheet.cell(row=i, column=7).value is None):
                break
            if sheet.cell(row=i, column=13).value is not None:
                existing_titles.append(str(sheet.cell(row=i, column=13).value).strip().lower())
            else:
                existing_titles.append(str(sheet.cell(row=i, column=4).value).strip().lower())
            i += 1
        w_row = i
        initial_row = w_row
        # Loop through publications of authors
        for index, row in data.iterrows():
            if (str(row['title'].strip().lower()) in existing_titles):
                continue
            sheet.cell(row=w_row, column=3).value = row['author'].replace(
                " and", ";")
            sheet.cell(row=w_row, column=4).value = row['title']
            sheet.cell(row=w_row, column=13).value = row['title']
            sheet.cell(row=w_row, column=12).value = row['abstract']
            if ('journal' in row):
                sheet.cell(row=w_row, column=5).value = row['journal']
            sheet.cell(row=w_row, column=6).hyperlink = str(row['pub_url'])
            if (not math.isnan(row['pub_year'])):
                sheet.cell(row=w_row, column=7).value = str(int(row['pub_year']))
            # Formatting
            sheet.cell(row=w_row, column=3).alignment = Alignment(
                wrap_text=True)
            sheet.cell(row=w_row, column=4).alignment = Alignment(
                wrap_text=True)
            sheet.cell(row=w_row, column=6).alignment = Alignment(
                wrap_text=True)
            sheet.cell(row=w_row, column=7).alignment = Alignment(
                horizontal='right')
            row_dim = sheet.row_dimensions[w_row]
            row_dim.height = 50
            w_row += 1
        final_row = w_row
        wb_obj.save(filename=path)
        wb_obj.close()
        notify = Email().send_email_notif(path, initial_row, final_row)
